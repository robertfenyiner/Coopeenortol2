"""
Servicio para generación de reportes financieros.
"""
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Optional, List
from io import BytesIO
from sqlalchemy.orm import Session
from sqlalchemy import func, and_

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models.asociado import Asociado
from app.models.contabilidad import CuentaContable, MovimientoContable, Aporte, AsientoContable
from app.models.credito import Credito
from app.models.ahorro import CuentaAhorro, MovimientoAhorro
from app.schemas.reportes import (
    BalanceGeneralResponse,
    GrupoBalance,
    CuentaBalance,
    EstadoResultadosResponse,
    CuentaResultados,
    ReporteCarteraResponse,
    CreditoCartera,
    EstadisticasCartera,
    ReporteMoraResponse,
    CreditoMora,
    EstadoCuentaAsociadoResponse,
    AportesAsociado,
    CreditoAsociado,
    CuentaAhorroAsociado,
    EstadisticasGeneralesResponse,
    EstadisticasAsociados,
    EstadisticasCarteraGeneral,
    EstadisticasAhorros
)


def generar_balance_general(db: Session, fecha_corte: date) -> BalanceGeneralResponse:
    """
    Generar Balance General a una fecha específica.
    """
    # Obtener todas las cuentas con movimientos hasta la fecha de corte
    cuentas = db.query(CuentaContable).filter(
        CuentaContable.activa == True
    ).all()
    
    activos = []
    pasivos = []
    patrimonio = []
    
    total_activos = Decimal("0.00")
    total_pasivos = Decimal("0.00")
    total_patrimonio = Decimal("0.00")
    
    for cuenta in cuentas:
        # Calcular saldo de la cuenta hasta la fecha de corte
        movimientos = db.query(MovimientoContable).join(
            AsientoContable, MovimientoContable.asiento_id == AsientoContable.id
        ).filter(
            MovimientoContable.cuenta_id == cuenta.id,
            AsientoContable.fecha <= fecha_corte,
            AsientoContable.anulado == False
        ).all()
        
        total_debito = sum(m.debito for m in movimientos)
        total_credito = sum(m.credito for m in movimientos)
        
        if cuenta.naturaleza == "debito":
            saldo = total_debito - total_credito
        else:
            saldo = total_credito - total_debito
        
        if saldo == 0:
            continue
        
        cuenta_balance = CuentaBalance(
            codigo=cuenta.codigo,
            nombre=cuenta.nombre,
            saldo=saldo
        )
        
        # Clasificar por tipo
        if cuenta.tipo == "activo":
            activos.append(cuenta_balance)
            total_activos += saldo
        elif cuenta.tipo == "pasivo":
            pasivos.append(cuenta_balance)
            total_pasivos += saldo
        elif cuenta.tipo == "patrimonio":
            patrimonio.append(cuenta_balance)
            total_patrimonio += saldo
    
    # Agrupar cuentas
    grupo_activos = [GrupoBalance(
        nombre="Activos",
        total=total_activos,
        cuentas=activos
    )]
    
    grupo_pasivos = [GrupoBalance(
        nombre="Pasivos",
        total=total_pasivos,
        cuentas=pasivos
    )]
    
    grupo_patrimonio = [GrupoBalance(
        nombre="Patrimonio",
        total=total_patrimonio,
        cuentas=patrimonio
    )]
    
    cuadrado = abs(total_activos - (total_pasivos + total_patrimonio)) < Decimal("0.01")
    
    return BalanceGeneralResponse(
        fecha_corte=fecha_corte,
        activos=grupo_activos,
        pasivos=grupo_pasivos,
        patrimonio=grupo_patrimonio,
        total_activos=total_activos,
        total_pasivos=total_pasivos,
        total_patrimonio=total_patrimonio,
        cuadrado=cuadrado
    )


def generar_estado_resultados(
    db: Session,
    fecha_inicio: date,
    fecha_fin: date
) -> EstadoResultadosResponse:
    """
    Generar Estado de Resultados para un período.
    """
    # Obtener cuentas de ingresos y gastos
    cuentas_ingresos = db.query(CuentaContable).filter(
        CuentaContable.tipo == "ingreso",
        CuentaContable.activa == True
    ).all()
    
    cuentas_gastos = db.query(CuentaContable).filter(
        CuentaContable.tipo == "gasto",
        CuentaContable.activa == True
    ).all()
    
    ingresos = []
    gastos = []
    total_ingresos = Decimal("0.00")
    total_gastos = Decimal("0.00")
    
    # Calcular ingresos
    for cuenta in cuentas_ingresos:
        movimientos = db.query(MovimientoContable).join(
            AsientoContable, MovimientoContable.asiento_id == AsientoContable.id
        ).filter(
            MovimientoContable.cuenta_id == cuenta.id,
            AsientoContable.fecha >= fecha_inicio,
            AsientoContable.fecha <= fecha_fin,
            AsientoContable.anulado == False
        ).all()
        
        total_credito = sum(m.credito for m in movimientos)
        total_debito = sum(m.debito for m in movimientos)
        valor = total_credito - total_debito
        
        if valor > 0:
            ingresos.append(CuentaResultados(
                codigo=cuenta.codigo,
                nombre=cuenta.nombre,
                valor=valor
            ))
            total_ingresos += valor
    
    # Calcular gastos
    for cuenta in cuentas_gastos:
        movimientos = db.query(MovimientoContable).join(
            AsientoContable, MovimientoContable.asiento_id == AsientoContable.id
        ).filter(
            MovimientoContable.cuenta_id == cuenta.id,
            AsientoContable.fecha >= fecha_inicio,
            AsientoContable.fecha <= fecha_fin,
            AsientoContable.anulado == False
        ).all()
        
        total_debito = sum(m.debito for m in movimientos)
        total_credito = sum(m.credito for m in movimientos)
        valor = total_debito - total_credito
        
        if valor > 0:
            gastos.append(CuentaResultados(
                codigo=cuenta.codigo,
                nombre=cuenta.nombre,
                valor=valor
            ))
            total_gastos += valor
    
    utilidad_neta = total_ingresos - total_gastos
    margen_utilidad = (utilidad_neta / total_ingresos * 100) if total_ingresos > 0 else Decimal("0.00")
    
    return EstadoResultadosResponse(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        ingresos=ingresos,
        gastos=gastos,
        total_ingresos=total_ingresos,
        total_gastos=total_gastos,
        utilidad_neta=utilidad_neta,
        margen_utilidad=margen_utilidad
    )


def generar_reporte_cartera(
    db: Session,
    fecha_corte: date,
    tipo_credito: Optional[str] = None,
    estado: Optional[str] = None
) -> ReporteCarteraResponse:
    """
    Generar reporte de cartera de créditos.
    """
    query = db.query(Credito).join(Asociado).filter(
        Credito.estado.in_(["desembolsado", "al_dia", "mora", "castigado"])
    )
    
    if tipo_credito:
        query = query.filter(Credito.tipo_credito == tipo_credito)
    
    if estado:
        query = query.filter(Credito.estado == estado)
    
    creditos_db = query.all()
    
    # Estadísticas
    total_creditos = len(creditos_db)
    cartera_total = sum(c.saldo_capital or Decimal("0.00") for c in creditos_db)
    cartera_al_dia = sum(
        c.saldo_capital or Decimal("0.00")
        for c in creditos_db
        if c.estado == "al_dia"
    )
    cartera_mora = sum(
        c.saldo_capital or Decimal("0.00")
        for c in creditos_db
        if c.estado == "mora"
    )
    cartera_castigada = sum(
        c.saldo_capital or Decimal("0.00")
        for c in creditos_db
        if c.estado == "castigado"
    )
    creditos_mora = len([c for c in creditos_db if c.estado == "mora"])
    tasa_mora = (cartera_mora / cartera_total * 100) if cartera_total > 0 else Decimal("0.00")
    monto_provision = cartera_mora * Decimal("0.05")  # 5% de provisión
    
    estadisticas = EstadisticasCartera(
        total_creditos=total_creditos,
        cartera_total=cartera_total,
        cartera_al_dia=cartera_al_dia,
        cartera_mora=cartera_mora,
        cartera_castigada=cartera_castigada,
        tasa_mora=tasa_mora,
        creditos_mora=creditos_mora,
        monto_provision=monto_provision
    )
    
    # Detalle de créditos
    creditos = []
    for c in creditos_db:
        creditos.append(CreditoCartera(
            numero_credito=c.numero_credito,
            asociado_nombre=f"{c.asociado.nombres} {c.asociado.apellidos}",
            asociado_documento=c.asociado.numero_documento,
            tipo_credito=c.tipo_credito,
            monto_desembolsado=c.monto_desembolsado or Decimal("0.00"),
            saldo_capital=c.saldo_capital or Decimal("0.00"),
            saldo_interes=c.saldo_interes or Decimal("0.00"),
            saldo_mora=c.saldo_mora or Decimal("0.00"),
            dias_mora=c.dias_mora or 0,
            estado=c.estado,
            fecha_desembolso=c.fecha_desembolso,
            fecha_ultimo_pago=c.fecha_ultimo_pago
        ))
    
    # Agrupar por tipo
    por_tipo = {}
    for c in creditos_db:
        if c.tipo_credito not in por_tipo:
            por_tipo[c.tipo_credito] = {
                "total": Decimal("0.00"),
                "creditos": 0
            }
        por_tipo[c.tipo_credito]["total"] += c.saldo_capital or Decimal("0.00")
        por_tipo[c.tipo_credito]["creditos"] += 1
    
    return ReporteCarteraResponse(
        fecha_corte=fecha_corte,
        estadisticas=estadisticas,
        creditos=creditos,
        por_tipo=por_tipo
    )


def generar_reporte_mora(db: Session, dias_mora_minimo: int) -> ReporteMoraResponse:
    """
    Generar reporte de créditos en mora.
    """
    creditos_db = db.query(Credito).join(Asociado).filter(
        Credito.estado == "mora",
        Credito.dias_mora >= dias_mora_minimo
    ).all()
    
    creditos = []
    monto_total_mora = Decimal("0.00")
    por_rango = {
        "1-30": {"creditos": 0, "monto": Decimal("0.00")},
        "31-60": {"creditos": 0, "monto": Decimal("0.00")},
        "61-90": {"creditos": 0, "monto": Decimal("0.00")},
        "90+": {"creditos": 0, "monto": Decimal("0.00")}
    }
    
    for c in creditos_db:
        dias = c.dias_mora or 0
        
        # Determinar rango
        if dias <= 30:
            rango = "1-30"
        elif dias <= 60:
            rango = "31-60"
        elif dias <= 90:
            rango = "61-90"
        else:
            rango = "90+"
        
        saldo_mora = c.saldo_mora or Decimal("0.00")
        monto_total_mora += saldo_mora
        por_rango[rango]["creditos"] += 1
        por_rango[rango]["monto"] += saldo_mora
        
        creditos.append(CreditoMora(
            numero_credito=c.numero_credito,
            asociado_id=c.asociado_id,
            asociado_nombre=f"{c.asociado.nombres} {c.asociado.apellidos}",
            asociado_documento=c.asociado.numero_documento,
            asociado_telefono=c.asociado.datos_personales.get("telefono"),
            tipo_credito=c.tipo_credito,
            saldo_capital=c.saldo_capital or Decimal("0.00"),
            saldo_mora=saldo_mora,
            dias_mora=dias,
            rango_mora=rango,
            fecha_ultimo_pago=c.fecha_ultimo_pago
        ))
    
    return ReporteMoraResponse(
        fecha_generacion=date.today(),
        dias_mora_minimo=dias_mora_minimo,
        total_creditos_mora=len(creditos),
        monto_total_mora=monto_total_mora,
        creditos=creditos,
        por_rango=por_rango
    )


def generar_estado_cuenta(
    db: Session,
    asociado_id: int,
    fecha_inicio: Optional[date],
    fecha_fin: date
) -> EstadoCuentaAsociadoResponse:
    """
    Generar estado de cuenta de un asociado.
    """
    if fecha_inicio is None:
        fecha_inicio = fecha_fin - timedelta(days=180)  # 6 meses atrás
    
    asociado = db.query(Asociado).filter(Asociado.id == asociado_id).first()
    if not asociado:
        raise ValueError("Asociado no encontrado")
    
    # Aportes
    aportes_db = db.query(Aporte).filter(
        Aporte.asociado_id == asociado_id,
        Aporte.fecha.between(fecha_inicio, fecha_fin)
    ).all()
    
    total_aportes = sum(a.valor for a in aportes_db)
    ultimo_aporte = max(aportes_db, key=lambda x: x.fecha) if aportes_db else None
    
    aportes = AportesAsociado(
        total_aportes=total_aportes,
        numero_aportes=len(aportes_db),
        ultimo_aporte_fecha=ultimo_aporte.fecha if ultimo_aporte else None,
        ultimo_aporte_valor=ultimo_aporte.valor if ultimo_aporte else None
    )
    
    # Créditos
    creditos_db = db.query(Credito).filter(
        Credito.asociado_id == asociado_id,
        Credito.estado.in_(["desembolsado", "al_dia", "mora"])
    ).all()
    
    creditos = []
    total_deuda = Decimal("0.00")
    for c in creditos_db:
        creditos.append(CreditoAsociado(
            numero_credito=c.numero_credito,
            tipo_credito=c.tipo_credito,
            monto_desembolsado=c.monto_desembolsado or Decimal("0.00"),
            saldo_capital=c.saldo_capital or Decimal("0.00"),
            valor_cuota=c.valor_cuota or Decimal("0.00"),
            estado=c.estado,
            dias_mora=c.dias_mora or 0
        ))
        total_deuda += c.saldo_capital or Decimal("0.00")
    
    # Cuentas de ahorro
    cuentas_db = db.query(CuentaAhorro).filter(
        CuentaAhorro.asociado_id == asociado_id,
        CuentaAhorro.estado == "activa"
    ).all()
    
    cuentas_ahorro = []
    total_ahorros = Decimal("0.00")
    for cuenta in cuentas_db:
        saldo_cuenta = cuenta.saldo_disponible + cuenta.saldo_bloqueado
        cuentas_ahorro.append(CuentaAhorroAsociado(
            numero_cuenta=cuenta.numero_cuenta,
            tipo_ahorro=cuenta.tipo_ahorro,
            saldo_actual=saldo_cuenta,
            estado=cuenta.estado
        ))
        total_ahorros += saldo_cuenta
    
    patrimonio_neto = total_aportes + total_ahorros - total_deuda
    
    return EstadoCuentaAsociadoResponse(
        asociado_id=asociado.id,
        nombres=asociado.nombres,
        apellidos=asociado.apellidos,
        numero_documento=asociado.numero_documento,
        fecha_generacion=date.today(),
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        aportes=aportes,
        creditos=creditos,
        cuentas_ahorro=cuentas_ahorro,
        total_aportes=total_aportes,
        total_deuda=total_deuda,
        total_ahorros=total_ahorros,
        patrimonio_neto=patrimonio_neto
    )


def generar_estado_cuenta_por_documento(
    db: Session,
    numero_documento: str,
    fecha_inicio: Optional[date],
    fecha_fin: date
) -> EstadoCuentaAsociadoResponse:
    """
    Generar estado de cuenta buscando por número de documento.
    """
    # Buscar asociado por número de documento
    asociado = db.query(Asociado).filter(
        Asociado.numero_documento == numero_documento
    ).first()
    
    if not asociado:
        raise ValueError(f"Asociado con documento {numero_documento} no encontrado")
    
    # Delegar a la función existente
    return generar_estado_cuenta(db, asociado.id, fecha_inicio, fecha_fin)


def obtener_estadisticas_generales(db: Session) -> EstadisticasGeneralesResponse:
    """
    Obtener estadísticas generales del sistema.
    """
    # Asociados
    total_asociados = db.query(Asociado).count()
    asociados_activos = db.query(Asociado).filter(Asociado.estado == "activo").count()
    asociados_inactivos = total_asociados - asociados_activos
    
    # Nuevos asociados este mes
    primer_dia_mes = date.today().replace(day=1)
    nuevos_mes = db.query(Asociado).filter(
        Asociado.created_at >= primer_dia_mes
    ).count()
    
    estadisticas_asociados = EstadisticasAsociados(
        total=total_asociados,
        activos=asociados_activos,
        inactivos=asociados_inactivos,
        nuevos_mes=nuevos_mes
    )
    
    # Cartera
    creditos_activos = db.query(Credito).filter(
        Credito.estado.in_(["desembolsado", "al_dia", "mora"])
    ).all()
    
    total_cartera = sum(c.saldo_capital or Decimal("0.00") for c in creditos_activos)
    cartera_al_dia = sum(
        c.saldo_capital or Decimal("0.00")
        for c in creditos_activos
        if c.estado in ["desembolsado", "al_dia"]
    )
    cartera_mora = sum(
        c.saldo_capital or Decimal("0.00")
        for c in creditos_activos
        if c.estado == "mora"
    )
    tasa_mora = (cartera_mora / total_cartera * 100) if total_cartera > 0 else Decimal("0.00")
    
    estadisticas_cartera = EstadisticasCarteraGeneral(
        total_cartera=total_cartera,
        cartera_al_dia=cartera_al_dia,
        cartera_mora=cartera_mora,
        tasa_mora=tasa_mora,
        creditos_activos=len(creditos_activos)
    )
    
    # Ahorros
    cuentas_activas = db.query(CuentaAhorro).filter(
        CuentaAhorro.estado == "activa"
    ).all()
    
    total_ahorros = sum((c.saldo_disponible + c.saldo_bloqueado) for c in cuentas_activas)
    ahorro_promedio = total_ahorros / len(cuentas_activas) if cuentas_activas else Decimal("0.00")
    
    estadisticas_ahorros = EstadisticasAhorros(
        total_ahorros=total_ahorros,
        cuentas_activas=len(cuentas_activas),
        ahorro_promedio=ahorro_promedio
    )
    
    # Aportes totales
    aportes_totales = db.query(func.sum(Aporte.valor)).scalar() or Decimal("0.00")
    
    # Operaciones del mes (simplificado)
    operaciones_mes = (
        db.query(Aporte).filter(Aporte.fecha >= primer_dia_mes).count() +
        db.query(Credito).filter(Credito.fecha_desembolso >= primer_dia_mes).count()
    )
    
    return EstadisticasGeneralesResponse(
        fecha_generacion=date.today(),
        asociados=estadisticas_asociados,
        cartera=estadisticas_cartera,
        ahorros=estadisticas_ahorros,
        aportes_totales=aportes_totales,
        operaciones_mes=operaciones_mes
    )


def exportar_balance_pdf(db: Session, fecha_corte: date) -> BytesIO:
    """
    Exportar balance general a PDF.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Generar datos del balance
    balance = generar_balance_general(db, fecha_corte)
    
    # Estilos personalizados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=6,
        alignment=TA_CENTER
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=12,
        spaceBefore=12
    )
    
    # Título
    elements.append(Paragraph("BALANCE GENERAL", title_style))
    elements.append(Paragraph(f"Al {fecha_corte.strftime('%d de %B de %Y')}", subtitle_style))
    
    # ACTIVOS
    elements.append(Paragraph("ACTIVOS", heading_style))
    activos_data = [['Código', 'Cuenta', 'Saldo']]
    
    for grupo in balance.activos:
        for cuenta in grupo.cuentas:
            activos_data.append([
                cuenta.codigo,
                cuenta.nombre,
                f"${cuenta.saldo:,.2f}"
            ])
    
    activos_data.append(['', 'TOTAL ACTIVOS', f"${balance.total_activos:,.2f}"])
    
    activos_table = Table(activos_data, colWidths=[1*inch, 4*inch, 1.5*inch])
    activos_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#dbeafe')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(activos_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # PASIVOS
    elements.append(Paragraph("PASIVOS", heading_style))
    pasivos_data = [['Código', 'Cuenta', 'Saldo']]
    
    for grupo in balance.pasivos:
        for cuenta in grupo.cuentas:
            pasivos_data.append([
                cuenta.codigo,
                cuenta.nombre,
                f"${cuenta.saldo:,.2f}"
            ])
    
    pasivos_data.append(['', 'TOTAL PASIVOS', f"${balance.total_pasivos:,.2f}"])
    
    pasivos_table = Table(pasivos_data, colWidths=[1*inch, 4*inch, 1.5*inch])
    pasivos_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ef4444')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fecaca')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(pasivos_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # PATRIMONIO
    elements.append(Paragraph("PATRIMONIO", heading_style))
    patrimonio_data = [['Código', 'Cuenta', 'Saldo']]
    
    for grupo in balance.patrimonio:
        for cuenta in grupo.cuentas:
            patrimonio_data.append([
                cuenta.codigo,
                cuenta.nombre,
                f"${cuenta.saldo:,.2f}"
            ])
    
    patrimonio_data.append(['', 'TOTAL PATRIMONIO', f"${balance.total_patrimonio:,.2f}"])
    
    patrimonio_table = Table(patrimonio_data, colWidths=[1*inch, 4*inch, 1.5*inch])
    patrimonio_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#a7f3d0')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(patrimonio_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Verificación de cuadre
    cuadre_ok = abs(balance.total_activos - (balance.total_pasivos + balance.total_patrimonio)) < 0.01
    cuadre_text = "✓ Balance Cuadrado" if cuadre_ok else "⚠ Balance Descuadrado"
    cuadre_color = colors.green if cuadre_ok else colors.red
    
    cuadre_style = ParagraphStyle(
        'Cuadre',
        parent=styles['Normal'],
        fontSize=12,
        textColor=cuadre_color,
        alignment=TA_CENTER,
        spaceAfter=12
    )
    elements.append(Paragraph(cuadre_text, cuadre_style))
    
    # Construir PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


def exportar_cartera_excel(db: Session, fecha_corte: date) -> BytesIO:
    """
    Exportar reporte de cartera a Excel.
    """
    # Generar datos de cartera
    reporte = generar_reporte_cartera(db, fecha_corte)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Cartera"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = f"REPORTE DE CARTERA AL {fecha_corte.strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Estadísticas
    ws['A3'] = "ESTADÍSTICAS"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A4'] = "Total Créditos:"
    ws['B4'] = reporte.estadisticas.total_creditos
    ws['A5'] = "Cartera Total:"
    ws['B5'] = reporte.estadisticas.cartera_total
    ws['B5'].number_format = '$#,##0.00'
    ws['A6'] = "Cartera al Día:"
    ws['B6'] = reporte.estadisticas.cartera_al_dia
    ws['B6'].number_format = '$#,##0.00'
    ws['A7'] = "Cartera en Mora:"
    ws['B7'] = reporte.estadisticas.cartera_mora
    ws['B7'].number_format = '$#,##0.00'
    ws['A8'] = "Tasa de Mora:"
    ws['B8'] = float(reporte.estadisticas.tasa_mora)
    ws['B8'].number_format = '0.00%'
    
    # Encabezados de tabla
    row = 11
    headers = ['ID', 'Asociado', 'Tipo', 'Monto Original', 'Saldo Actual', 'Cuota', 'Estado', 'Días Mora']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    row += 1
    for credito in reporte.creditos:
        ws.cell(row=row, column=1, value=credito.numero_credito).border = border
        ws.cell(row=row, column=2, value=credito.asociado_nombre).border = border
        ws.cell(row=row, column=3, value=credito.tipo_credito).border = border
        
        cell = ws.cell(row=row, column=4, value=float(credito.monto_desembolsado))
        cell.number_format = '$#,##0.00'
        cell.border = border
        
        cell = ws.cell(row=row, column=5, value=float(credito.saldo_capital))
        cell.number_format = '$#,##0.00'
        cell.border = border
        
        cell = ws.cell(row=row, column=6, value=0)  # No hay cuota en el schema
        cell.number_format = '$#,##0.00'
        cell.border = border
        
        ws.cell(row=row, column=7, value=credito.estado).border = border
        ws.cell(row=row, column=8, value=credito.dias_mora).border = border
        row += 1
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def exportar_estado_resultados_pdf(db: Session, fecha_inicio: date, fecha_fin: date) -> BytesIO:
    """
    Exportar Estado de Resultados a PDF.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Generar datos del estado
    estado = generar_estado_resultados(db, fecha_inicio, fecha_fin)
    
    # Estilos
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, 
                                  textColor=colors.HexColor('#1e40af'), spaceAfter=6, alignment=TA_CENTER)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, 
                                     textColor=colors.grey, spaceAfter=20, alignment=TA_CENTER)
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=14, 
                                    textColor=colors.HexColor('#1e40af'), spaceAfter=12, spaceBefore=12)
    
    # Título
    elements.append(Paragraph("ESTADO DE RESULTADOS", title_style))
    elements.append(Paragraph(
        f"Del {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}", 
        subtitle_style
    ))
    
    # KPIs
    kpi_data = [
        ['Total Ingresos', f"${estado.total_ingresos:,.2f}"],
        ['Total Gastos', f"${estado.total_gastos:,.2f}"],
        ['Utilidad Neta', f"${estado.utilidad_neta:,.2f}"],
        ['Margen Utilidad', f"{estado.margen_utilidad:.2f}%"]
    ]
    kpi_table = Table(kpi_data, colWidths=[3*inch, 2*inch])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e0e7ff')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # INGRESOS
    elements.append(Paragraph("INGRESOS", heading_style))
    
    if estado.ingresos:
        ingresos_data = [['Código', 'Cuenta', 'Valor']]
        for ing in estado.ingresos:
            ingresos_data.append([ing.codigo, ing.nombre, f"${ing.valor:,.2f}"])
        
        ingresos_table = Table(ingresos_data, colWidths=[1*inch, 4*inch, 1.5*inch])
        ingresos_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(ingresos_table)
        elements.append(Spacer(1, 0.2*inch))
    
    total_ingresos_data = [['TOTAL INGRESOS', f"${estado.total_ingresos:,.2f}"]]
    total_ingresos_table = Table(total_ingresos_data, colWidths=[4.5*inch, 1.5*inch])
    total_ingresos_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#a7f3d0')),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
    ]))
    elements.append(total_ingresos_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # GASTOS
    elements.append(Paragraph("GASTOS", heading_style))
    
    if estado.gastos:
        gastos_data = [['Código', 'Cuenta', 'Valor']]
        for gasto in estado.gastos:
            gastos_data.append([gasto.codigo, gasto.nombre, f"${gasto.valor:,.2f}"])
        
        gastos_table = Table(gastos_data, colWidths=[1*inch, 4*inch, 1.5*inch])
        gastos_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ef4444')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(gastos_table)
        elements.append(Spacer(1, 0.2*inch))
    
    total_gastos_data = [['TOTAL GASTOS', f"${estado.total_gastos:,.2f}"]]
    total_gastos_table = Table(total_gastos_data, colWidths=[4.5*inch, 1.5*inch])
    total_gastos_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#fecaca')),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
    ]))
    elements.append(total_gastos_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # RESULTADOS
    resultados_data = [
        ['Total Ingresos', f"${estado.total_ingresos:,.2f}"],
        ['Total Gastos', f"${estado.total_gastos:,.2f}"],
        ['UTILIDAD NETA', f"${estado.utilidad_neta:,.2f}"],
        ['Margen Utilidad', f"{estado.margen_utilidad:.2f}%"]
    ]
    resultados_table = Table(resultados_data, colWidths=[4.5*inch, 1.5*inch])
    resultados_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -2), colors.HexColor('#dbeafe')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(resultados_table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def exportar_mora_excel(db: Session, dias_mora_minimo: int = 1) -> BytesIO:
    """
    Exportar Reporte de Mora a Excel.
    """
    # Generar datos de mora
    reporte = generar_reporte_mora(db, dias_mora_minimo)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Mora"
    
    # Estilos
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:I1')
    ws['A1'] = f"REPORTE DE MOROSIDAD - {reporte.fecha_generacion.strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Estadísticas
    ws['A3'] = "Total Créditos en Mora:"
    ws['B3'] = reporte.total_creditos_mora
    ws['A4'] = "Monto Total Mora:"
    ws['B4'] = float(reporte.monto_total_mora)
    ws['B4'].number_format = '$#,##0.00'
    ws['A5'] = "Días Mínimos:"
    ws['B5'] = reporte.dias_mora_minimo
    
    # Encabezados
    row = 8
    headers = ['Número Crédito', 'Asociado', 'Documento', 'Teléfono', 'Tipo Crédito', 
               'Saldo Capital', 'Saldo Mora', 'Días Mora', 'Rango']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    row += 1
    for credito in reporte.creditos:
        ws.cell(row=row, column=1, value=credito.numero_credito).border = border
        ws.cell(row=row, column=2, value=credito.asociado_nombre).border = border
        ws.cell(row=row, column=3, value=credito.asociado_documento).border = border
        ws.cell(row=row, column=4, value=credito.asociado_telefono or '').border = border
        ws.cell(row=row, column=5, value=credito.tipo_credito).border = border
        
        cell = ws.cell(row=row, column=6, value=float(credito.saldo_capital))
        cell.number_format = '$#,##0.00'
        cell.border = border
        
        cell = ws.cell(row=row, column=7, value=float(credito.saldo_mora))
        cell.number_format = '$#,##0.00'
        cell.border = border
        
        ws.cell(row=row, column=8, value=credito.dias_mora).border = border
        ws.cell(row=row, column=9, value=credito.rango_mora).border = border
        
        row += 1
    
    # Ajustar anchos
    for col, width in [(1, 12), (2, 30), (3, 15), (4, 25), (5, 15), (6, 12), (7, 15), (8, 15), (9, 15)]:
        ws.column_dimensions[get_column_letter(col)].width = width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def exportar_estado_cuenta_excel(db: Session, asociado_id: int, fecha_inicio: Optional[date], 
                                  fecha_fin: date) -> BytesIO:
    """
    Exportar Estado de Cuenta del Asociado a Excel con formato profesional y logo.
    """
    from openpyxl.drawing.image import Image as XLImage
    import os
    
    # Generar datos del estado de cuenta
    estado = generar_estado_cuenta(db, asociado_id, fecha_inicio, fecha_fin)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de Cuenta"
    
    # Colores corporativos
    color_header = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    color_subheader = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
    color_section = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
    color_total = PatternFill(start_color="93c5fd", end_color="93c5fd", fill_type="solid")
    
    font_white = Font(color="FFFFFF", bold=True, size=12)
    font_bold = Font(bold=True, size=11)
    font_title = Font(bold=True, size=16, color="1e40af")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Agregar logo si existe
    logo_path = "/root/projects/Coopeenortol/assets/logos/logo-principal.png"
    if os.path.exists(logo_path):
        img = XLImage(logo_path)
        # Ajustar tamaño del logo a 80x80 px
        img.width = 80
        img.height = 80
        ws.add_image(img, 'A1')
        # Ajustar altura de filas para el logo
        for i in range(1, 6):
            ws.row_dimensions[i].height = 15
        row = 7
    
    # Título
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws.cell(row=row, column=1, value="ESTADO DE CUENTA")
    cell.font = font_title
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    # Subtítulo con nombre del asociado
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws.cell(row=row, column=1, value=f"{estado.nombres} {estado.apellidos}")
    cell.font = Font(size=12, bold=True)
    cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # Fecha de generación
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws.cell(row=row, column=1, value=f"Generado el {estado.fecha_generacion.strftime('%d de %B de %Y')}")
    cell.font = Font(size=10, italic=True, color="666666")
    cell.alignment = Alignment(horizontal='center')
    row += 2
    
    # INFORMACIÓN DEL ASOCIADO
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws.cell(row=row, column=1, value="INFORMACIÓN DEL ASOCIADO")
    cell.font = font_white
    cell.fill = color_header
    cell.alignment = Alignment(horizontal='center')
    cell.border = border
    row += 1
    
    info_data = [
        ('ID Asociado:', str(estado.asociado_id)),
        ('Documento:', estado.numero_documento),
        ('Periodo:', f"{estado.fecha_inicio.strftime('%d/%m/%Y') if estado.fecha_inicio else 'Inicio'} - {estado.fecha_fin.strftime('%d/%m/%Y')}")
    ]
    
    for label, value in info_data:
        ws.cell(row=row, column=1, value=label).font = font_bold
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2, value=value).border = border
        ws.merge_cells(f'B{row}:F{row}')
        row += 1
    
    row += 1
    
    # RESUMEN FINANCIERO
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws.cell(row=row, column=1, value="RESUMEN FINANCIERO")
    cell.font = font_white
    cell.fill = color_header
    cell.alignment = Alignment(horizontal='center')
    cell.border = border
    row += 1
    
    # KPIs
    kpi_data = [
        ('Total Aportes', float(estado.total_aportes)),
        ('Total Deuda', float(estado.total_deuda)),
        ('Total Ahorros', float(estado.total_ahorros)),
        ('Patrimonio Neto', float(estado.patrimonio_neto))
    ]
    
    for label, value in kpi_data:
        ws.cell(row=row, column=1, value=label).font = font_bold
        ws.cell(row=row, column=1).fill = color_section
        ws.cell(row=row, column=1).border = border
        
        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = '$#,##0.00'
        cell.font = Font(size=11)
        cell.alignment = Alignment(horizontal='right')
        cell.border = border
        ws.merge_cells(f'B{row}:F{row}')
        row += 1
    
    row += 1
    
    # RESUMEN DE APORTES
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws.cell(row=row, column=1, value="RESUMEN DE APORTES")
    cell.font = font_white
    cell.fill = color_subheader
    cell.alignment = Alignment(horizontal='center')
    cell.border = border
    row += 1
    
    aportes_data = [
        ('Total Aportes', float(estado.aportes.total_aportes)),
        ('Número de Aportes', estado.aportes.numero_aportes),
        ('Último Aporte', float(estado.aportes.ultimo_aporte_valor) if estado.aportes.ultimo_aporte_valor else 0),
        ('Fecha Último Aporte', estado.aportes.ultimo_aporte_fecha.strftime('%d/%m/%Y') if estado.aportes.ultimo_aporte_fecha else 'N/A')
    ]
    
    for label, value in aportes_data:
        ws.cell(row=row, column=1, value=label).font = font_bold
        ws.cell(row=row, column=1).border = border
        
        cell = ws.cell(row=row, column=2, value=value)
        if label != 'Fecha Último Aporte' and label != 'Número de Aportes':
            cell.number_format = '$#,##0.00'
        cell.border = border
        cell.alignment = Alignment(horizontal='right')
        ws.merge_cells(f'B{row}:F{row}')
        row += 1
    
    row += 1
    
    # CRÉDITOS
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws.cell(row=row, column=1, value="CRÉDITOS")
    cell.font = font_white
    cell.fill = color_subheader
    cell.alignment = Alignment(horizontal='center')
    cell.border = border
    row += 1
    
    if estado.creditos:
        # Encabezados de tabla
        headers = ['Número', 'Tipo', 'Desembolsado', 'Saldo Capital', 'Cuota', 'Días Mora', 'Estado']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = font_white
            cell.fill = color_header
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        row += 1
        
        # Datos de créditos
        for credito in estado.creditos:
            ws.cell(row=row, column=1, value=credito.numero_credito).border = border
            ws.cell(row=row, column=2, value=credito.tipo_credito).border = border
            
            cell = ws.cell(row=row, column=3, value=float(credito.monto_desembolsado))
            cell.number_format = '$#,##0.00'
            cell.border = border
            
            cell = ws.cell(row=row, column=4, value=float(credito.saldo_capital))
            cell.number_format = '$#,##0.00'
            cell.border = border
            
            cell = ws.cell(row=row, column=5, value=float(credito.valor_cuota))
            cell.number_format = '$#,##0.00'
            cell.border = border
            
            ws.cell(row=row, column=6, value=credito.dias_mora).border = border
            ws.cell(row=row, column=7, value=credito.estado).border = border
            row += 1
    else:
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws.cell(row=row, column=1, value="No hay créditos activos")
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        row += 1
    
    row += 1
    
    # CUENTAS DE AHORRO
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws.cell(row=row, column=1, value="CUENTAS DE AHORRO")
    cell.font = font_white
    cell.fill = color_subheader
    cell.alignment = Alignment(horizontal='center')
    cell.border = border
    row += 1
    
    if estado.cuentas_ahorro:
        # Encabezados
        headers = ['Número Cuenta', 'Tipo', 'Saldo', 'Estado']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = font_white
            cell.fill = color_header
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        row += 1
        
        # Datos de cuentas
        for cuenta in estado.cuentas_ahorro:
            ws.cell(row=row, column=1, value=cuenta.numero_cuenta).border = border
            ws.cell(row=row, column=2, value=cuenta.tipo_ahorro).border = border
            
            cell = ws.cell(row=row, column=3, value=float(cuenta.saldo_actual))
            cell.number_format = '$#,##0.00'
            cell.border = border
            
            ws.cell(row=row, column=4, value=cuenta.estado).border = border
            row += 1
    else:
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws.cell(row=row, column=1, value="No hay cuentas de ahorro")
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        row += 1
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def exportar_estado_cuenta_excel_por_documento(
    db: Session,
    numero_documento: str,
    fecha_inicio: Optional[date],
    fecha_fin: date
) -> BytesIO:
    """
    Exportar Estado de Cuenta a Excel usando número de documento.
    """
    asociado = db.query(Asociado).filter(Asociado.numero_documento == numero_documento).first()
    if not asociado:
        raise ValueError(f"Asociado con documento {numero_documento} no encontrado")
    return exportar_estado_cuenta_excel(db, asociado.id, fecha_inicio, fecha_fin)


def exportar_estado_cuenta_pdf(db: Session, asociado_id: int, fecha_inicio: Optional[date], 
                                fecha_fin: date) -> BytesIO:
    """
    Exportar Estado de Cuenta del Asociado a PDF con logo.
    """
    import os
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Generar datos del estado de cuenta
    estado = generar_estado_cuenta(db, asociado_id, fecha_inicio, fecha_fin)
    
    # Estilos
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, 
                                  textColor=colors.HexColor('#1e40af'), spaceAfter=6, alignment=TA_CENTER)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, 
                                     textColor=colors.grey, spaceAfter=20, alignment=TA_CENTER)
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=14, 
                                    textColor=colors.HexColor('#1e40af'), spaceAfter=12, spaceBefore=12)
    
    # Agregar logo si existe
    logo_path = "/root/projects/Coopeenortol/assets/logos/logo-principal.png"
    if os.path.exists(logo_path):
        logo = RLImage(logo_path, width=1*inch, height=1*inch)
        elements.append(logo)
        elements.append(Spacer(1, 0.2*inch))
    
    # Título
    elements.append(Paragraph("ESTADO DE CUENTA", title_style))
    elements.append(Paragraph(f"Asociado: {estado.nombres} {estado.apellidos}", subtitle_style))
    elements.append(Paragraph(f"Al {estado.fecha_generacion.strftime('%d de %B de %Y')}", subtitle_style))
    
    # Información del Asociado
    info_data = [
        ['ID Asociado:', str(estado.asociado_id)],
        ['Documento:', estado.numero_documento],
        ['Periodo:', f"{estado.fecha_inicio.strftime('%d/%m/%Y') if estado.fecha_inicio else 'Inicio'} - {estado.fecha_fin.strftime('%d/%m/%Y')}"]
    ]
    info_table = Table(info_data, colWidths=[2*inch, 3*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e0e7ff')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # KPIs Financieros
    kpi_data = [
        ['Total Aportes', f"${estado.total_aportes:,.2f}"],
        ['Total Deuda', f"${estado.total_deuda:,.2f}"],
        ['Total Ahorros', f"${estado.total_ahorros:,.2f}"],
        ['Patrimonio Neto', f"${estado.patrimonio_neto:,.2f}"]
    ]
    kpi_table = Table(kpi_data, colWidths=[3*inch, 2*inch])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#dbeafe')),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # APORTES
    elements.append(Paragraph("RESUMEN DE APORTES", heading_style))
    aportes_data = [
        ['Total Aportes', f"${estado.aportes.total_aportes:,.2f}"],
        ['Número de Aportes', str(estado.aportes.numero_aportes)],
        ['Último Aporte', f"${estado.aportes.ultimo_aporte_valor:,.2f}" if estado.aportes.ultimo_aporte_valor else "N/A"],
        ['Fecha Último Aporte', estado.aportes.ultimo_aporte_fecha.strftime('%d/%m/%Y') if estado.aportes.ultimo_aporte_fecha else "N/A"]
    ]
    aportes_table = Table(aportes_data, colWidths=[4*inch, 2*inch])
    aportes_table.setStyle(TableStyle([
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(aportes_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # CRÉDITOS
    if estado.creditos:
        elements.append(Paragraph("CRÉDITOS", heading_style))
        creditos_data = [['Número', 'Tipo', 'Desembolso', 'Saldo', 'Cuota', 'Estado', 'Mora']]
        for credito in estado.creditos:
            creditos_data.append([
                credito.numero_credito,
                credito.tipo_credito,
                f"${credito.monto_desembolsado:,.2f}",
                f"${credito.saldo_capital:,.2f}",
                f"${credito.valor_cuota:,.2f}",
                credito.estado,
                str(credito.dias_mora)
            ])
        
        creditos_table = Table(creditos_data, colWidths=[1*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1*inch, 0.9*inch, 0.6*inch])
        creditos_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ef4444')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 1), (-2, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(creditos_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # CUENTAS DE AHORRO
    if estado.cuentas_ahorro:
        elements.append(Paragraph("CUENTAS DE AHORRO", heading_style))
        ahorros_data = [['Número Cuenta', 'Tipo', 'Saldo', 'Estado']]
        for cuenta in estado.cuentas_ahorro:
            ahorros_data.append([
                cuenta.numero_cuenta,
                cuenta.tipo_ahorro,
                f"${cuenta.saldo_actual:,.2f}",
                cuenta.estado
            ])
        
        ahorros_table = Table(ahorros_data, colWidths=[1.5*inch, 2.5*inch, 1.5*inch, 1*inch])
        ahorros_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 1), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(ahorros_table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def exportar_estado_cuenta_pdf_por_documento(
    db: Session,
    numero_documento: str,
    fecha_inicio: Optional[date],
    fecha_fin: date
) -> BytesIO:
    """
    Exportar Estado de Cuenta a PDF buscando por número de documento.
    """
    # Buscar asociado por número de documento
    asociado = db.query(Asociado).filter(
        Asociado.numero_documento == numero_documento
    ).first()
    
    if not asociado:
        raise ValueError(f"Asociado con documento {numero_documento} no encontrado")
    
    # Delegar a la función existente
    return exportar_estado_cuenta_pdf(db, asociado.id, fecha_inicio, fecha_fin)


def generar_certificado_paz_salvo(
    db: Session,
    numero_documento: str
) -> BytesIO:
    """
    Generar certificado de Paz y Salvo para un asociado.
    
    Certifica que el asociado no tiene obligaciones pendientes.
    """
    # Buscar asociado
    asociado = db.query(Asociado).filter(
        Asociado.numero_documento == numero_documento
    ).first()
    
    if not asociado:
        raise ValueError(f"Asociado con documento {numero_documento} no encontrado")
    
    # Verificar obligaciones
    creditos_activos = db.query(Credito).filter(
        Credito.asociado_id == asociado.id,
        Credito.estado.in_(["desembolsado", "al_dia", "mora"])
    ).all()
    
    tiene_deudas = any(c.saldo_capital > 0 for c in creditos_activos)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilo del título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e3a8a'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Encabezado
    elements.append(Spacer(1, 0.5*inch))
    elements.append(Paragraph("COOPERATIVA COOPEENORTOL", title_style))
    elements.append(Paragraph("NIT: 900.XXX.XXX-X", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Título del certificado
    cert_title = ParagraphStyle(
        'CertTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#dc2626') if tiene_deudas else colors.HexColor('#059669'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    if tiene_deudas:
        elements.append(Paragraph("CERTIFICADO DE ESTADO DE OBLIGACIONES", cert_title))
    else:
        elements.append(Paragraph("CERTIFICADO DE PAZ Y SALVO", cert_title))
    
    elements.append(Spacer(1, 0.3*inch))
    
    # Contenido
    content_style = ParagraphStyle(
        'Content',
        parent=styles['Normal'],
        fontSize=11,
        leading=18,
        alignment=TA_LEFT
    )
    
    fecha_actual = datetime.now().strftime("%d de %B de %Y")
    
    texto = f"""
    <para align=justify>
    La Cooperativa COOPEENORTOL certifica que el(la) asociado(a) 
    <b>{asociado.nombres} {asociado.apellidos}</b>, identificado(a) con 
    cédula de ciudadanía No. <b>{asociado.numero_documento}</b>, 
    {'' if tiene_deudas else 'se encuentra a <b>PAZ Y SALVO</b> por todo concepto con esta cooperativa.'}
    </para>
    """
    
    elements.append(Paragraph(texto, content_style))
    elements.append(Spacer(1, 0.2*inch))
    
    if tiene_deudas:
        # Mostrar obligaciones pendientes
        elements.append(Paragraph("<b>Obligaciones Pendientes:</b>", content_style))
        elements.append(Spacer(1, 0.1*inch))
        
        deudas_data = [['Número Crédito', 'Tipo', 'Saldo Capital', 'Saldo Total', 'Estado']]
        total_deuda = Decimal("0.00")
        
        for credito in creditos_activos:
            if credito.saldo_capital > 0:
                saldo_total = (credito.saldo_capital or Decimal("0.00")) + \
                             (credito.saldo_interes or Decimal("0.00")) + \
                             (credito.saldo_mora or Decimal("0.00"))
                total_deuda += saldo_total
                
                deudas_data.append([
                    credito.numero_credito,
                    credito.tipo_credito,
                    f"${credito.saldo_capital:,.2f}",
                    f"${saldo_total:,.2f}",
                    credito.estado
                ])
        
        deudas_data.append(['', '', '', f"<b>${total_deuda:,.2f}</b>", ''])
        
        deudas_table = Table(deudas_data, colWidths=[1.5*inch, 1.5*inch, 1.2*inch, 1.2*inch, 1*inch])
        deudas_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (2, 1), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fee2e2')),
        ]))
        elements.append(deudas_table)
        elements.append(Spacer(1, 0.2*inch))
        
        elements.append(Paragraph(
            "El asociado NO se encuentra a paz y salvo hasta cancelar las obligaciones mencionadas.",
            content_style
        ))
    else:
        elements.append(Paragraph(
            "El presente certificado se expide a solicitud del interesado para los fines que estime convenientes.",
            content_style
        ))
    
    elements.append(Spacer(1, 0.5*inch))
    elements.append(Paragraph(f"Fecha de expedición: {fecha_actual}", content_style))
    elements.append(Spacer(1, inch))
    
    # Firmas
    firma_data = [
        ['_' * 30, '', '_' * 30],
        ['Gerente General', '', 'Contador']
    ]
    
    firma_table = Table(firma_data, colWidths=[2*inch, 1*inch, 2*inch])
    firma_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    elements.append(firma_table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generar_certificado_aportes(
    db: Session,
    numero_documento: str,
    ano: Optional[int] = None
) -> BytesIO:
    """
    Generar certificado de aportes de un asociado.
    
    Certifica el total de aportes realizados por el asociado.
    """
    # Buscar asociado
    asociado = db.query(Asociado).filter(
        Asociado.numero_documento == numero_documento
    ).first()
    
    if not asociado:
        raise ValueError(f"Asociado con documento {numero_documento} no encontrado")
    
    # Obtener aportes
    query = db.query(Aporte).filter(Aporte.asociado_id == asociado.id)
    
    if ano:
        fecha_inicio = date(ano, 1, 1)
        fecha_fin = date(ano, 12, 31)
        query = query.filter(
            Aporte.fecha_aporte >= fecha_inicio,
            Aporte.fecha_aporte <= fecha_fin
        )
    
    aportes = query.all()
    total_aportes = sum(a.valor for a in aportes)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilo del título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e3a8a'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Encabezado
    elements.append(Spacer(1, 0.5*inch))
    elements.append(Paragraph("COOPERATIVA COOPEENORTOL", title_style))
    elements.append(Paragraph("NIT: 900.XXX.XXX-X", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Título del certificado
    cert_title = ParagraphStyle(
        'CertTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#059669'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    periodo_texto = f"AÑO {ano}" if ano else "HISTÓRICO"
    elements.append(Paragraph(f"CERTIFICADO DE APORTES - {periodo_texto}", cert_title))
    elements.append(Spacer(1, 0.3*inch))
    
    # Contenido
    content_style = ParagraphStyle(
        'Content',
        parent=styles['Normal'],
        fontSize=11,
        leading=18,
        alignment=TA_LEFT
    )
    
    fecha_actual = datetime.now().strftime("%d de %B de %Y")
    
    texto = f"""
    <para align=justify>
    La Cooperativa COOPEENORTOL certifica que el(la) asociado(a) 
    <b>{asociado.nombres} {asociado.apellidos}</b>, identificado(a) con 
    cédula de ciudadanía No. <b>{asociado.numero_documento}</b>, ha realizado 
    aportes a la cooperativa {f'durante el año {ano}' if ano else 'desde su ingreso'} 
    por un valor total de:
    </para>
    """
    
    elements.append(Paragraph(texto, content_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Total en recuadro destacado
    total_style = ParagraphStyle(
        'Total',
        parent=styles['Normal'],
        fontSize=24,
        textColor=colors.HexColor('#059669'),
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    elements.append(Paragraph(f"${total_aportes:,.2f}", total_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Detalle si hay aportes
    if aportes:
        elements.append(Paragraph("<b>Detalle de Aportes:</b>", content_style))
        elements.append(Spacer(1, 0.1*inch))
        
        aportes_data = [['Fecha', 'Valor', 'Tipo', 'Referencia']]
        for aporte in aportes[-10:]:  # Últimos 10 aportes
            aportes_data.append([
                aporte.fecha_aporte.strftime("%Y-%m-%d"),
                f"${aporte.valor:,.2f}",
                aporte.tipo_aporte,
                aporte.referencia or "N/A"
            ])
        
        if len(aportes) > 10:
            aportes_data.append(['...', '...', '...', '...'])
        
        aportes_table = Table(aportes_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 2*inch])
        aportes_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(aportes_table)
        elements.append(Spacer(1, 0.2*inch))
        
        if len(aportes) > 10:
            elements.append(Paragraph(
                f"(Mostrando los últimos 10 de {len(aportes)} aportes registrados)",
                content_style
            ))
    
    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph(
        "El presente certificado se expide a solicitud del interesado para los fines que estime convenientes.",
        content_style
    ))
    
    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph(f"Fecha de expedición: {fecha_actual}", content_style))
    elements.append(Spacer(1, inch))
    
    # Firmas
    firma_data = [
        ['_' * 30, '', '_' * 30],
        ['Gerente General', '', 'Contador']
    ]
    
    firma_table = Table(firma_data, colWidths=[2*inch, 1*inch, 2*inch])
    firma_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    elements.append(firma_table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer
